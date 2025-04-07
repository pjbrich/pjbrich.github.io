#created by Benjamin Richards 04/04/2025
#still need to have this script read through the execl file to find all titanlinks and
import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime

def scrape_product_info(url):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Referer': 'https://www.titanfittings.com/'
        }

        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Try multiple selectors to find the product title
        selectors = [
            'h1.product-name',                # Original selector
            'h1.product-title',               # Common alternative
            'h1.product-detail__title',       # Another common pattern
            'h1.productView-title',           # Yet another common pattern
            'h1[itemprop="name"]',           # Schema.org markup
            'h1.title'                        # Simple fallback
        ]

        product_title = None
        for selector in selectors:
            element = soup.select_one(selector)
            if element:
                product_title = element.get_text(strip=True)
                break

        if not product_title:
            # Fallback to searching for any h1 if specific selectors fail
            h1 = soup.find('h1')
            if h1:
                product_title = h1.get_text(strip=True)

        if not product_title:
            raise ValueError("Could not find product title on page")

        return product_title

    except requests.exceptions.RequestException as e:
        print(f"Error during request to {url}: {str(e)}")
        return None
    except ValueError as e:
        print(f"Error parsing content from {url}: {str(e)}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while processing {url}: {str(e)}")
        return None

def save_to_excel(data, file_path):
    try:
        # Try to load existing workbook
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        except FileNotFoundError:
            # If file doesn't exist, create new workbook
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Product Information", "Date Added", "Product URL"])  # Add headers with URL
        except Exception as e:
            print(f"Error loading/creating Excel file {file_path}: {e}")
            return

        # Write data
        sheet.append([data, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), product_url])

        # Save workbook
        wb.save(file_path)
        print(f"Data successfully saved to {file_path}")

    except Exception as e:
        print(f"Error saving to Excel: {e}")

if __name__ == "__main__":
    # Define the path to the Excel file containing the links
    links_excel_path = "BOP.xlsx"  # Assuming BOP.xlsx is in the same directory
    output_excel_path = "BOP_output.xlsx" # Name for the output file

    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(links_excel_path)
        sheet = workbook.active

        # Iterate through the rows in column J (index 9, as columns are 0-indexed)
        for row in range(1, sheet.max_row + 1):  # Start from row 1 (assuming headers are in the first row)
            cell_value = sheet.cell(row=row, column=10).value  # Column J is the 10th column

            if cell_value and isinstance(cell_value, str) and cell_value.startswith("https://www.titanfittings.com/"):
                product_url = cell_value
                print(f"Processing URL: {product_url}")
                product_info = scrape_product_info(product_url)

                if product_info:
                    print(f"Scraped product info: {product_info}")
                    save_to_excel(product_info, output_excel_path)
                else:
                    print(f"Failed to scrape product information for {product_url}")
            elif cell_value:
                print(f"Skipping non-Titan Fittings link or empty cell in row {row}, column J: {cell_value}")
            else:
                print(f"Skipping empty cell in row {row}, column J.")

        print("Finished processing all links.")

    except FileNotFoundError:
        print(f"Error: The file '{links_excel_path}' was not found in the current directory.")
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")